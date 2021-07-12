import requests
from time import sleep
from bs4 import BeautifulSoup
from openpyxl import Workbook


core_url = 'https://www.ihk-lehrstellenboerse.de'
wb = Workbook()

dest_filename = 'iwex-v3.xlsx'

ws = wb.active
ws.title = 'Information'
counter = 1
for i in range(2080, 2827):
    shared_url = f'https://www.ihk-lehrstellenboerse.de/angebote/suche?hitsPerPage=10&page={i}&sortColumn=-1&sortDir=asc&query=Gib+Deinen+Wunschberuf+ein&organisationName=Unternehmen+eingeben&status=1&mode=0&dateTypeSelection=LASTCHANGED_DATE&thisYear=true&nextYear=true&afterNextYear=true&distance=0'
    response = requests.get(shared_url)
    soup = BeautifulSoup(response.text, 'html.parser')
    table = soup.find('table', {'class': 'sortableTable'})
    tbody = table.find('tbody')
    trs = tbody.find_all('tr')
    for tr in trs:
        # try:
        td = tr.find('td')
        a = td.find('a')
        url = a['href']
        response = requests.get(f'{core_url}{url}')
        print(f'{core_url}{url}')
        soup = BeautifulSoup(response.text, 'html.parser')
        div = soup.find('div', {'class': 'contentBox clearfix contactBox'})
        name = div.findChild('h3')
        address = div.findChild('p')
        a = address.text
        a = a.replace('\t', '').replace('\n', '')
        ps = div.findChildren('p')
        person = ''
        telephone = ''
        email = ''

        h3 = soup.find('h3', {'class': 'mt20'})
        if h3:
            ps = h3.fetchNextSiblings()

            for p in ps:
                if p is ps[0]:
                    person = p.text
                elif p.text.startswith('Tel.'):
                    telephone = p.text[5:].strip()
                elif p.text.startswith('E-Mail:'):
                    email = p.text[7:].strip()

        ws[f'A{counter}'] = name.text
        ws[f'B{counter}'] = a
        ws[f'C{counter}'] = person
        ws[f'D{counter}'] = telephone
        ws[f'E{counter}'] = email
        counter += 1
        print(counter)
        # except:
        #     continue

wb.save(filename=dest_filename)

